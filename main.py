# coding:utf-8
import openpyxl

x = 2

w1 = openpyxl.load_workbook('1.xlsx')
w2 = openpyxl.load_workbook('3.xlsx')

# 新建record的xlsx文件
w3 = openpyxl.Workbook()
sheet3 = w3.active
sheet3.title = 'record'

a = w1.sheetnames
b = w2.sheetnames

sheet1 = w1.get_sheet_by_name(a[0])
sheet2 = w2.get_sheet_by_name(b[0])

max1 = sheet1.max_row
max2 = sheet2.max_row

m1 = sheet1.max_column
m2 = sheet2.max_column

# print('123')
# print(sheet2.cell(row=2,column=1).value)
#
for i in range(2,max2 + 1):
    for j in range(2,max1 + 1):
        if sheet1.cell(row=j,column=1).value == sheet2.cell(row=i,column=1).value:
            for k in range(1,m1+1):
                sheet3.cell(row=x,column=k).value = sheet1.cell(row=j,column=k).value
            x += 1

w3.save('record.xlsx')





